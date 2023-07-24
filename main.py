import json
from openpyxl import load_workbook


tb = load_workbook('sample.xlsx')
tb_names = tb.sheetnames

keys = ['C', 'M', 'Y', 'K', 'Pantone']  # ключи полей

pantone = []


def parser():
    for item in tb_names:
        page = tb[item]
        for col in page.iter_cols():
            obj = []
            for cell in col:
                data_str = cell.value
                if data_str and not ("PANTONE" in data_str):
                    for a in data_str.split(" "):
                        obj.append(int(a.split(":")[1]))
                elif data_str:
                    obj.append(data_str)
                if len(obj) == 5:
                    obj.append(obj.pop(0))
                    pantone.append(dict(zip(keys, obj)))
                    print(obj)
                    obj = []

    with open('data.json', 'w') as data_f:
        json.dump(pantone, data_f)


if __name__ == "__main__":
    parser()
